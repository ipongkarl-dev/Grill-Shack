from fastapi import FastAPI, APIRouter, HTTPException, Query, Request
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import bcrypt
import jwt as pyjwt

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI(title="Grill Shack Management API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ===================== AUTH HELPERS =====================

JWT_ALGORITHM = "HS256"

def get_jwt_secret():
    return os.environ.get("JWT_SECRET", "fallback-secret-key")

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {"sub": user_id, "email": email, "role": role, "exp": datetime.now(timezone.utc) + timedelta(hours=24), "type": "access"}
    return pyjwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return pyjwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = pyjwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user_id = payload["sub"]
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user.pop("password_hash", None)
        return user
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except pyjwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


async def require_owner(request: Request) -> dict:
    """Get current user and verify they have owner role."""
    user = await get_current_user(request)
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Owner access required")
    return user

async def seed_admin():
    admin_email = os.environ.get("ADMIN_EMAIL", "owner@grillshack.nz").lower()
    admin_password = os.environ.get("ADMIN_PASSWORD", "GrillShack2026!")
    existing = await db.users.find_one({"email": admin_email}, {"_id": 0})
    if not existing:
        user_id = str(uuid.uuid4())
        await db.users.insert_one({
            "id": user_id, "email": admin_email, "name": "Owner",
            "password_hash": hash_password(admin_password),
            "role": "owner", "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Admin seeded: {admin_email}")
    elif not verify_password(admin_password, existing.get("password_hash", "")):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})
    # Write test credentials
    try:
        creds_path = Path("/app/memory/test_credentials.md")
        creds_path.parent.mkdir(exist_ok=True)
        creds_path.write_text(f"# Auth Credentials\n## Owner\n- Email: {admin_email}\n- Password: {admin_password}\n- Role: owner\n\n## Staff (create via app)\n- Role: staff\n")
    except Exception:
        pass



# ===================== MODELS =====================

class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    code: str
    unit: str = "pcs"
    price: float
    food_cost: float
    packaging_cost: float
    total_cost: float = 0
    cogs_percent: float = 0
    profit_per_order: float = 0
    opening_stock: int = 0
    current_stock: int = 0
    reorder_point: int = 10
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

    def __init__(self, **data):
        super().__init__(**data)
        self.total_cost = self.food_cost + self.packaging_cost
        if self.price > 0:
            self.cogs_percent = (self.total_cost / self.price) * 100
            self.profit_per_order = self.price - self.total_cost

class ProductCreate(BaseModel):
    name: str
    code: str
    unit: str = "pcs"
    price: float
    food_cost: float
    packaging_cost: float
    opening_stock: int = 0
    reorder_point: int = 10

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    unit: Optional[str] = None
    price: Optional[float] = None
    food_cost: Optional[float] = None
    packaging_cost: Optional[float] = None
    opening_stock: Optional[int] = None
    current_stock: Optional[int] = None
    reorder_point: Optional[int] = None

class Market(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    preset_mix: Dict[str, float] = {}
    is_active: bool = True

class MarketCreate(BaseModel):
    name: str
    preset_mix: Dict[str, float] = {}

class SessionSale(BaseModel):
    product_id: str
    product_name: str
    units_sold: int
    unit_price: float
    sales_value: float = 0
    food_cogs: float = 0
    packaging: float = 0
    total_cogs: float = 0
    gross_profit: float = 0

class Session(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    session_id: str
    date: str
    market_id: str
    market_name: str
    cash: float = 0
    eftpos: float = 0
    total_collected: float = 0
    sales: List[SessionSale] = []
    total_units: int = 0
    calculated_sales: float = 0
    variance: float = 0
    status: str = "OK"
    food_cogs: float = 0
    packaging: float = 0
    total_cogs: float = 0
    gross_profit: float = 0
    cogs_percent: float = 0
    opening_float: float = 0
    cash_expenses: float = 0
    expense_notes: str = ""
    notes: str = ""
    created_by_id: str = ""
    created_by_name: str = ""
    created_by_role: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SessionCreate(BaseModel):
    date: str
    market_id: str
    market_name: str
    cash: float = 0
    eftpos: float = 0
    sales: List[Dict[str, Any]]
    opening_float: float = 0
    cash_expenses: float = 0
    expense_notes: str = ""
    notes: str = ""
    created_by_id: str = ""
    created_by_name: str = ""
    created_by_role: str = ""

class AllocationSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "default"
    owner_pay_percent: float = 30
    growth_percent: float = 20
    emergency_percent: float = 10
    buffer_percent: float = 40
    gst_rate: float = 15

class AllocationSettingsUpdate(BaseModel):
    owner_pay_percent: Optional[float] = None
    growth_percent: Optional[float] = None
    emergency_percent: Optional[float] = None
    buffer_percent: Optional[float] = None
    gst_rate: Optional[float] = None

class InventoryEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    ingredient_name: str
    date: str
    packs_in: int = 0
    units_per_pack: float = 1
    unit: str = "kg"
    total_qty_added: float = 0
    pack_cost: float = 0
    cost_per_unit: float = 0
    supplier: str = ""
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InventoryEntryCreate(BaseModel):
    ingredient_name: str
    date: str
    packs_in: int = 0
    units_per_pack: float = 1
    unit: str = "kg"
    pack_cost: float = 0
    supplier: str = ""
    notes: str = ""

class CashflowTarget(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    month: str  # YYYY-MM
    sales_target: float = 0
    actual_sales: float = 0
    growth_saved: float = 0
    emergency_saved: float = 0
    notes: str = ""

class CashflowTargetCreate(BaseModel):
    month: str
    sales_target: float = 0
    growth_saved: float = 0
    emergency_saved: float = 0
    notes: str = ""

class ProductIngredient(BaseModel):
    name: str
    qty_per_order: float = 0
    unit: str = "g"
    cost_per_unit: float = 0
    total_cost: float = 0

class ProductWithIngredients(BaseModel):
    model_config = ConfigDict(extra="ignore")
    product_id: str
    ingredients: List[Dict[str, Any]] = []

class ProductIngredientsUpdate(BaseModel):
    ingredients: List[Dict[str, Any]]

# ===================== HELPER FUNCTIONS =====================

def serialize_doc(doc: dict) -> dict:
    """Remove MongoDB _id and serialize datetime fields"""
    if doc is None:  # noqa: E711 - intentional None check
        return None
    doc.pop('_id', None)
    for key, value in doc.items():
        if isinstance(value, datetime):
            doc[key] = value.isoformat()
    return doc

async def get_next_session_id():
    """Generate next session ID"""
    last_session = await db.sessions.find_one({}, sort=[("session_id", -1)])
    if last_session and "session_id" in last_session:
        try:
            last_num = int(last_session["session_id"])
            return str(last_num + 1)
        except (ValueError, TypeError, IndexError):
            pass
    return "46130"


def process_session_sales(sales_input: list, products_map: dict):
    """Extract sales processing logic from create_session. Returns (session_sales, totals)."""
    session_sales = []
    totals = {"units": 0, "sales": 0, "food_cogs": 0, "packaging": 0}

    for sale in sales_input:
        product_id = sale.get('product_id')
        units = sale.get('units_sold', 0)
        if product_id not in products_map or units <= 0:
            continue
        product = products_map[product_id]
        sales_value = units * product['price']
        item_food_cogs = units * product['food_cost']
        item_packaging = units * product['packaging_cost']
        item_total_cogs = item_food_cogs + item_packaging

        session_sales.append(SessionSale(
            product_id=product_id,
            product_name=product['name'],
            units_sold=units,
            unit_price=product['price'],
            sales_value=sales_value,
            food_cogs=item_food_cogs,
            packaging=item_packaging,
            total_cogs=item_total_cogs,
            gross_profit=sales_value - item_total_cogs
        ))
        totals["units"] += units
        totals["sales"] += sales_value
        totals["food_cogs"] += item_food_cogs
        totals["packaging"] += item_packaging

    return session_sales, totals

def determine_cash_status(calculated_sales: float, total_collected: float) -> str:
    """Determine cash reconciliation status."""
    if total_collected == 0 and calculated_sales > 0:
        return "Missing Payment"
    variance = calculated_sales - total_collected
    if variance > 0.5:
        return "Under-collected"
    if variance < -0.5:
        return "Over-collected"
    return "OK"


def build_session_financials(totals: dict, cash: float, eftpos: float):
    """Calculate session financial fields from sales totals and payment info."""
    total_cogs = totals["food_cogs"] + totals["packaging"]
    calculated_sales = totals["sales"]
    gross_profit = calculated_sales - total_cogs
    cogs_percent = (total_cogs / calculated_sales * 100) if calculated_sales > 0 else 0
    total_collected = cash + eftpos
    variance = calculated_sales - total_collected
    status = determine_cash_status(calculated_sales, total_collected)
    return {
        "total_cogs": round(total_cogs, 2),
        "calculated_sales": round(calculated_sales, 2),
        "gross_profit": round(gross_profit, 2),
        "cogs_percent": round(cogs_percent, 2),
        "total_collected": total_collected,
        "variance": round(variance, 2),
        "status": status,
        "food_cogs": round(totals["food_cogs"], 2),
        "packaging": round(totals["packaging"], 2),
        "total_units": totals["units"],
    }


def compute_weekly_metrics(sessions):
    """Compute weekly activity metrics from sessions."""
    from datetime import date as dt_date
    week_set = set()
    total_sales = sum(s.get('calculated_sales', 0) for s in sessions)
    total_cogs = sum(s.get('total_cogs', 0) for s in sessions)
    for s in sessions:
        try:
            parts = s.get('date', '').split('-')
            d = dt_date(int(parts[0]), int(parts[1]), int(parts[2]))
            week_set.add(d.isocalendar()[:2])
        except (ValueError, TypeError, IndexError):
            pass
    weeks_active = max(len(week_set), 1)
    return {
        "total_sales": total_sales,
        "total_cogs": total_cogs,
        "weeks_active": weeks_active,
        "current_weekly": total_sales / weeks_active,
        "avg_cogs_pct": (total_cogs / total_sales) if total_sales > 0 else 0.25,
    }


def compute_scale_projections(target_weekly_revenue, weeks_horizon, avg_session_rev,
                               total_sessions, weekly_metrics, products):
    """Compute growth projections for scale planner."""
    gst_rate_val = weekly_metrics.get("gst_rate", 0.15)
    avg_cogs_pct = weekly_metrics["avg_cogs_pct"]
    current_weekly = weekly_metrics["current_weekly"]
    weeks_active = weekly_metrics["weeks_active"]

    sessions_per_week_needed = target_weekly_revenue / avg_session_rev if avg_session_rev > 0 else 4
    sessions_per_week_current = total_sessions / weeks_active

    projected_gross = target_weekly_revenue * weeks_horizon
    projected_net = projected_gross / (1 + gst_rate_val)
    projected_cogs = projected_net * avg_cogs_pct
    projected_profit = projected_net - projected_cogs

    total_stock_value = sum(p.get('current_stock', 0) * p.get('food_cost', 0) for p in products)
    weekly_stock_needed = target_weekly_revenue * avg_cogs_pct

    revenue_gap = target_weekly_revenue - current_weekly
    growth_pct = (revenue_gap / current_weekly * 100) if current_weekly > 0 else 0

    return {
        "sessions_per_week_current": round(sessions_per_week_current, 1),
        "sessions_per_week_needed": round(sessions_per_week_needed, 1),
        "revenue_gap": round(revenue_gap, 2),
        "growth_needed_pct": round(growth_pct, 1),
        "projections": {
            "gross_revenue": round(projected_gross, 2),
            "net_revenue": round(projected_net, 2),
            "total_cogs": round(projected_cogs, 2),
            "total_profit": round(projected_profit, 2),
        },
        "investment": {
            "current_stock_value": round(total_stock_value, 2),
            "weekly_stock_investment": round(weekly_stock_needed, 2),
            "total_stock_investment": round(weekly_stock_needed * weeks_horizon, 2),
        },
    }

def aggregate_by_period(sessions: list, key_fn) -> dict:
    """Generic aggregation helper for sessions by a period key function."""
    period_data = {}
    for s in sessions:
        key = key_fn(s)
        if not key:
            continue
        if key not in period_data:
            period_data[key] = {'sales': 0, 'profit': 0, 'units': 0, 'sessions': 0, 'cogs': 0}
        period_data[key]['sales'] += s.get('calculated_sales', 0)
        period_data[key]['profit'] += s.get('gross_profit', 0)
        period_data[key]['units'] += s.get('total_units', 0)
        period_data[key]['sessions'] += 1
        period_data[key]['cogs'] += s.get('total_cogs', 0)
    return period_data


# ===================== ROUTES =====================

@api_router.get("/")
async def root():
    return {"message": "Grill Shack Management API", "version": "1.0"}

# -------- PRODUCTS --------

@api_router.get("/products", response_model=List[Product])
async def get_products():
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    return products

@api_router.get("/products/{product_id}", response_model=Product)
async def get_product(product_id: str):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product

@api_router.post("/products", response_model=Product)
async def create_product(input: ProductCreate, request: Request):
    await require_owner(request)
    product = Product(
        name=input.name,
        code=input.code,
        unit=input.unit,
        price=input.price,
        food_cost=input.food_cost,
        packaging_cost=input.packaging_cost,
        opening_stock=input.opening_stock,
        current_stock=input.opening_stock,
        reorder_point=input.reorder_point
    )
    doc = product.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.products.insert_one(doc)
    return product

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, input: ProductUpdate, request: Request):
    await require_owner(request)
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    # Recalculate derived fields if cost/price changes
    existing = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    
    price = update_data.get('price', existing['price'])
    food_cost = update_data.get('food_cost', existing['food_cost'])
    packaging_cost = update_data.get('packaging_cost', existing['packaging_cost'])
    
    update_data['total_cost'] = food_cost + packaging_cost
    if price > 0:
        update_data['cogs_percent'] = (update_data['total_cost'] / price) * 100
        update_data['profit_per_order'] = price - update_data['total_cost']
    
    await db.products.update_one({"id": product_id}, {"$set": update_data})
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    return updated

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, request: Request):
    await require_owner(request)
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted"}

# -------- MARKETS --------

@api_router.get("/markets", response_model=List[Market])
async def get_markets():
    markets = await db.markets.find({}, {"_id": 0}).to_list(100)
    return markets

@api_router.post("/markets", response_model=Market)
async def create_market(input: MarketCreate, request: Request):
    await require_owner(request)
    market = Market(name=input.name, preset_mix=input.preset_mix)
    await db.markets.insert_one(market.model_dump())
    return market

@api_router.put("/markets/{market_id}")
async def update_market(market_id: str, input: MarketCreate, request: Request):
    await require_owner(request)
    result = await db.markets.update_one(
        {"id": market_id},
        {"$set": {"name": input.name, "preset_mix": input.preset_mix}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Market not found")
    return {"message": "Market updated"}

@api_router.delete("/markets/{market_id}")
async def delete_market(market_id: str, request: Request):
    await require_owner(request)
    result = await db.markets.delete_one({"id": market_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Market not found")
    return {"message": "Market deleted"}

# -------- SESSIONS --------

@api_router.get("/sessions", response_model=List[Session])
async def get_sessions(
    market_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = Query(100, le=500)
):
    query = {}
    if market_id:
        query["market_id"] = market_id
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    sessions = await db.sessions.find(query, {"_id": 0}).sort("date", -1).to_list(limit)
    return sessions

@api_router.get("/sessions/{session_id}", response_model=Session)
async def get_session(session_id: str):
    session = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    return session

@api_router.post("/sessions", response_model=Session)
async def create_session(input: SessionCreate):
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    products_map = {p['id']: p for p in products}

    session_sales, totals = process_session_sales(input.sales, products_map)

    # Update product stock for each sold item
    for sale in session_sales:
        product = products_map.get(sale.product_id)
        if product:
            new_stock = max(0, product['current_stock'] - sale.units_sold)
            await db.products.update_one({"id": sale.product_id}, {"$set": {"current_stock": new_stock}})

    fin = build_session_financials(totals, input.cash, input.eftpos)
    session_id = await get_next_session_id()

    session = Session(
        session_id=session_id,
        date=input.date,
        market_id=input.market_id,
        market_name=input.market_name,
        cash=input.cash,
        eftpos=input.eftpos,
        total_collected=fin["total_collected"],
        sales=[s.model_dump() for s in session_sales],
        total_units=fin["total_units"],
        calculated_sales=fin["calculated_sales"],
        variance=fin["variance"],
        status=fin["status"],
        food_cogs=fin["food_cogs"],
        packaging=fin["packaging"],
        total_cogs=fin["total_cogs"],
        gross_profit=fin["gross_profit"],
        cogs_percent=fin["cogs_percent"],
        opening_float=input.opening_float,
        cash_expenses=input.cash_expenses,
        expense_notes=input.expense_notes,
        notes=input.notes,
        created_by_id=input.created_by_id,
        created_by_name=input.created_by_name,
        created_by_role=input.created_by_role
    )

    doc = session.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.sessions.insert_one(doc)

    return session

@api_router.put("/sessions/{session_id}")
async def update_session(session_id: str, input: Dict[str, Any]):
    existing = await db.sessions.find_one({"id": session_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Session not found")

    # If sales are being re-entered, recalculate everything
    if 'sales' in input:
        products = await db.products.find({}, {"_id": 0}).to_list(100)
        products_map = {p['id']: p for p in products}
        session_sales, totals = process_session_sales(input['sales'], products_map)
        cash = input.get('cash', existing.get('cash', 0))
        eftpos = input.get('eftpos', existing.get('eftpos', 0))
        fin = build_session_financials(totals, cash, eftpos)

        await db.sessions.update_one({"id": session_id}, {"$set": {
            "date": input.get('date', existing.get('date')),
            "market_id": input.get('market_id', existing.get('market_id')),
            "market_name": input.get('market_name', existing.get('market_name')),
            "cash": cash, "eftpos": eftpos,
            "total_collected": fin["total_collected"],
            "sales": [s.model_dump() for s in session_sales],
            "total_units": fin["total_units"],
            "calculated_sales": fin["calculated_sales"],
            "variance": fin["variance"],
            "status": fin["status"],
            "food_cogs": fin["food_cogs"],
            "packaging": fin["packaging"],
            "total_cogs": fin["total_cogs"],
            "gross_profit": fin["gross_profit"],
            "cogs_percent": fin["cogs_percent"],
            "opening_float": input.get('opening_float', existing.get('opening_float', 0)),
            "cash_expenses": input.get('cash_expenses', existing.get('cash_expenses', 0)),
            "expense_notes": input.get('expense_notes', existing.get('expense_notes', '')),
            "notes": input.get('notes', existing.get('notes', '')),
        }})
        return {"message": "Session fully updated with recalculated COGS"}

    # Otherwise just update allowed fields
    allowed_fields = ['cash', 'eftpos', 'opening_float', 'cash_expenses', 'expense_notes', 'notes', 'date', 'market_id', 'market_name']
    update_data = {k: v for k, v in input.items() if k in allowed_fields}

    if 'cash' in update_data or 'eftpos' in update_data:
        cash = update_data.get('cash', existing.get('cash', 0))
        eftpos = update_data.get('eftpos', existing.get('eftpos', 0))
        total_collected = cash + eftpos
        calculated_sales = existing.get('calculated_sales', 0)
        variance = calculated_sales - total_collected
        update_data['total_collected'] = total_collected
        update_data['variance'] = round(variance, 2)
        update_data['status'] = determine_cash_status(calculated_sales, total_collected)

    await db.sessions.update_one({"id": session_id}, {"$set": update_data})
    return {"message": "Session updated"}

@api_router.delete("/sessions/{session_id}")
async def delete_session(session_id: str):
    result = await db.sessions.delete_one({"id": session_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"message": "Session deleted"}

# -------- DASHBOARD & ANALYTICS --------

@api_router.get("/dashboard/kpis")
async def get_dashboard_kpis():
    # Get all sessions
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    
    total_sales = sum(s.get('calculated_sales', 0) for s in sessions)
    total_cogs = sum(s.get('total_cogs', 0) for s in sessions)
    total_profit = sum(s.get('gross_profit', 0) for s in sessions)
    total_units = sum(s.get('total_units', 0) for s in sessions)
    session_count = len(sessions)
    
    avg_cogs_percent = (total_cogs / total_sales * 100) if total_sales > 0 else 0
    
    # Low stock alerts
    low_stock_products = [p for p in products if p.get('current_stock', 0) <= p.get('reorder_point', 10)]
    
    # Recent sessions (last 7)
    recent_sessions = sorted(sessions, key=lambda x: x.get('date', ''), reverse=True)[:7]
    
    # Sales by market
    market_sales = {}
    total_cash_expenses = 0
    for s in sessions:
        market = s.get('market_name', 'Unknown')
        if market not in market_sales:
            market_sales[market] = {'sales': 0, 'units': 0, 'sessions': 0}
        market_sales[market]['sales'] += s.get('calculated_sales', 0)
        market_sales[market]['units'] += s.get('total_units', 0)
        market_sales[market]['sessions'] += 1
        total_cash_expenses += s.get('cash_expenses', 0)

    # Net profit = gross profit - cash expenses (market fees, overheads)
    settings_doc = await db.allocation_settings.find_one({"id": "default"}, {"_id": 0})
    gst_rate = (settings_doc.get('gst_rate', 15) if settings_doc else 15) / 100
    gst_amount = total_sales * gst_rate / (1 + gst_rate)
    net_profit = total_profit - total_cash_expenses - gst_amount

    return {
        "total_sales": round(total_sales, 2),
        "total_cogs": round(total_cogs, 2),
        "total_profit": round(total_profit, 2),
        "net_profit": round(net_profit, 2),
        "gst_amount": round(gst_amount, 2),
        "total_cash_expenses": round(total_cash_expenses, 2),
        "total_units": total_units,
        "session_count": session_count,
        "avg_cogs_percent": round(avg_cogs_percent, 2),
        "low_stock_products": low_stock_products,
        "recent_sessions": recent_sessions,
        "market_sales": market_sales
    }

@api_router.get("/dashboard/sales-by-month")
async def get_sales_by_month():
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    
    monthly_data = {}
    for s in sessions:
        date_str = s.get('date', '')
        if date_str:
            month = date_str[:7]  # YYYY-MM
            if month not in monthly_data:
                monthly_data[month] = {'sales': 0, 'cogs': 0, 'profit': 0, 'units': 0}
            monthly_data[month]['sales'] += s.get('calculated_sales', 0)
            monthly_data[month]['cogs'] += s.get('total_cogs', 0)
            monthly_data[month]['profit'] += s.get('gross_profit', 0)
            monthly_data[month]['units'] += s.get('total_units', 0)
    
    result = [
        {"month": k, **{key: round(val, 2) for key, val in v.items()}}
        for k, v in sorted(monthly_data.items())
    ]
    return result

@api_router.get("/dashboard/product-performance")
async def get_product_performance():
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    
    product_stats = {}
    for p in products:
        product_stats[p['id']] = {
            'name': p['name'],
            'code': p['code'],
            'price': p['price'],
            'cogs_percent': p.get('cogs_percent', 0),
            'units_sold': 0,
            'revenue': 0,
            'cogs': 0,
            'profit': 0
        }
    
    for session in sessions:
        for sale in session.get('sales', []):
            pid = sale.get('product_id')
            if pid in product_stats:
                product_stats[pid]['units_sold'] += sale.get('units_sold', 0)
                product_stats[pid]['revenue'] += sale.get('sales_value', 0)
                product_stats[pid]['cogs'] += sale.get('total_cogs', 0)
                product_stats[pid]['profit'] += sale.get('gross_profit', 0)
    
    result = list(product_stats.values())
    result.sort(key=lambda x: x['revenue'], reverse=True)
    return result

@api_router.get("/dashboard/margin-watch")
async def get_margin_watch():
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    
    # Calculate lifetime stats per product
    product_stats = {}
    for p in products:
        product_stats[p['id']] = {
            'id': p['id'],
            'name': p['name'],
            'code': p['code'],
            'price': p['price'],
            'unit_cost': p.get('total_cost', 0),
            'cogs_percent': p.get('cogs_percent', 0),
            'profit_per_order': p.get('profit_per_order', 0),
            'current_stock': p.get('current_stock', 0),
            'lifetime_units': 0,
            'lifetime_revenue': 0,
            'lifetime_profit': 0,
            'action': 'KEEP VISIBLE'
        }
    
    for session in sessions:
        for sale in session.get('sales', []):
            pid = sale.get('product_id')
            if pid in product_stats:
                product_stats[pid]['lifetime_units'] += sale.get('units_sold', 0)
                product_stats[pid]['lifetime_revenue'] += sale.get('sales_value', 0)
                product_stats[pid]['lifetime_profit'] += sale.get('gross_profit', 0)
    
    # Determine action recommendations
    for pid, stats in product_stats.items():
        if stats['cogs_percent'] > 35:
            stats['action'] = 'CHECK PRICE'
        elif stats['cogs_percent'] < 25 and stats['lifetime_units'] > 50:
            stats['action'] = 'PUSH HARD'
        elif stats['lifetime_units'] < 10:
            stats['action'] = 'PROMOTE'
    
    result = list(product_stats.values())
    result.sort(key=lambda x: x['lifetime_revenue'], reverse=True)
    return result

# -------- STOCK PLANNER --------

@api_router.get("/stock-planner")
async def get_stock_planner(market_id: Optional[str] = None, target_revenue: float = 1000):
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    
    # Calculate average sales mix from historical data
    product_totals = {}
    total_units = 0
    
    market_sessions = sessions
    if market_id:
        market_sessions = [s for s in sessions if s.get('market_id') == market_id]
    
    for session in market_sessions:
        for sale in session.get('sales', []):
            pid = sale.get('product_id')
            units = sale.get('units_sold', 0)
            if pid not in product_totals:
                product_totals[pid] = 0
            product_totals[pid] += units
            total_units += units
    
    # Build stock plan
    stock_plan = []
    for p in products:
        avg_mix = (product_totals.get(p['id'], 0) / total_units * 100) if total_units > 0 else 10
        estimated_orders = int((target_revenue * (avg_mix / 100)) / p['price']) if p['price'] > 0 else 0
        current_stock = p.get('current_stock', 0)
        gap = max(0, estimated_orders - current_stock)
        
        stock_plan.append({
            'product_id': p['id'],
            'product_name': p['name'],
            'code': p['code'],
            'sales_mix_percent': round(avg_mix, 2),
            'estimated_orders': estimated_orders,
            'current_stock': current_stock,
            'gap_to_buy': gap,
            'status': 'Covered' if gap == 0 else 'Buy/Prep',
            'unit_cost': p.get('total_cost', 0),
            'estimated_cogs': round(estimated_orders * p.get('total_cost', 0), 2),
            'stock_alert': 'LOW' if current_stock <= p.get('reorder_point', 10) else 'OK'
        })
    
    return {
        "target_revenue": target_revenue,
        "market_id": market_id,
        "stock_plan": stock_plan
    }

# -------- ALLOCATION TOOL --------

@api_router.get("/allocation/settings", response_model=AllocationSettings)
async def get_allocation_settings():
    settings = await db.allocation_settings.find_one({"id": "default"}, {"_id": 0})
    if not settings:
        default = AllocationSettings()
        await db.allocation_settings.insert_one(default.model_dump())
        return default
    return settings

@api_router.put("/allocation/settings", response_model=AllocationSettings)
async def update_allocation_settings(input: AllocationSettingsUpdate, request: Request):
    await require_owner(request)
    update_data = {k: v for k, v in input.model_dump().items() if v is not None}
    await db.allocation_settings.update_one(
        {"id": "default"},
        {"$set": update_data},
        upsert=True
    )
    return await get_allocation_settings()

@api_router.get("/allocation/calculate")
async def calculate_allocation(week_sales: float = 0):
    settings = await get_allocation_settings()
    
    # Calculate GST (inclusive)
    gst_rate = settings.get('gst_rate', 15) / 100
    gst_amount = week_sales * gst_rate / (1 + gst_rate)
    net_sales = week_sales - gst_amount
    
    # Get average COGS % from recent sessions
    sessions = await db.sessions.find({}, {"_id": 0}).sort("date", -1).to_list(10)
    if sessions:
        total_sales = sum(s.get('calculated_sales', 0) for s in sessions)
        total_cogs = sum(s.get('total_cogs', 0) for s in sessions)
        avg_cogs_percent = (total_cogs / total_sales) if total_sales > 0 else 0.25
    else:
        avg_cogs_percent = 0.25
    
    estimated_cogs = net_sales * avg_cogs_percent
    gross_profit = net_sales - estimated_cogs
    
    # Allocations
    owner_pay = gross_profit * (settings.get('owner_pay_percent', 30) / 100)
    growth = gross_profit * (settings.get('growth_percent', 20) / 100)
    emergency = gross_profit * (settings.get('emergency_percent', 10) / 100)
    buffer = gross_profit * (settings.get('buffer_percent', 40) / 100)
    
    return {
        "week_sales": round(week_sales, 2),
        "gst_amount": round(gst_amount, 2),
        "net_sales": round(net_sales, 2),
        "estimated_cogs": round(estimated_cogs, 2),
        "gross_profit": round(gross_profit, 2),
        "allocations": {
            "owner_pay": round(owner_pay, 2),
            "growth": round(growth, 2),
            "emergency": round(emergency, 2),
            "buffer": round(buffer, 2)
        },
        "settings": settings if isinstance(settings, dict) else settings.model_dump()
    }

# -------- INVENTORY --------

@api_router.get("/inventory")
async def get_inventory_entries(ingredient_name: Optional[str] = None):
    query = {}
    if ingredient_name:
        query["ingredient_name"] = ingredient_name
    entries = await db.inventory.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return entries

async def cascade_ingredient_cost(ingredient_name: str, cost_per_unit: float):
    """Update all product ingredients matching this name and recalculate COGS."""
    all_pi = await db.product_ingredients.find({}, {"_id": 0}).to_list(500)
    for pi_doc in all_pi:
        updated = False
        for ing in pi_doc.get('ingredients', []):
            if ing.get('name', '').lower() == ingredient_name.lower():
                ing['cost_per_unit'] = round(cost_per_unit, 4)
                ing['total_cost'] = round(ing.get('qty_per_order', 0) * cost_per_unit, 4)
                updated = True
        if not updated:
            continue
        total_food_cost = round(sum(i.get('total_cost', 0) for i in pi_doc['ingredients']), 2)
        await db.product_ingredients.update_one(
            {"product_id": pi_doc['product_id']},
            {"$set": {"ingredients": pi_doc['ingredients']}}
        )
        product = await db.products.find_one({"id": pi_doc['product_id']}, {"_id": 0})
        if product:
            pkg = product.get('packaging_cost', 0)
            total_cost = total_food_cost + pkg
            price = product.get('price', 0)
            cogs_pct = (total_cost / price * 100) if price > 0 else 0
            await db.products.update_one({"id": pi_doc['product_id']}, {"$set": {
                "food_cost": total_food_cost, "total_cost": round(total_cost, 2),
                "cogs_percent": round(cogs_pct, 2), "profit_per_order": round(price - total_cost, 2)
            }})

@api_router.post("/inventory")
async def create_inventory_entry(input: InventoryEntryCreate):
    total_qty = input.packs_in * input.units_per_pack
    cost_per_unit = (input.pack_cost / input.units_per_pack) if input.units_per_pack > 0 else 0

    entry = InventoryEntry(
        ingredient_name=input.ingredient_name, date=input.date,
        packs_in=input.packs_in, units_per_pack=input.units_per_pack,
        unit=input.unit, total_qty_added=round(total_qty, 4),
        pack_cost=input.pack_cost, cost_per_unit=round(cost_per_unit, 4),
        supplier=input.supplier, notes=input.notes
    )
    doc = entry.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.inventory.insert_one(doc)
    await cascade_ingredient_cost(input.ingredient_name, cost_per_unit)
    return entry

# -------- UNIQUE INGREDIENTS LIST --------

@api_router.get("/inventory/ingredients")
async def get_unique_ingredients():
    """Get list of unique ingredient names from inventory history"""
    entries = await db.inventory.find({}, {"_id": 0, "ingredient_name": 1, "unit": 1, "cost_per_unit": 1, "supplier": 1, "date": 1}).sort("date", -1).to_list(1000)
    # Get latest entry per ingredient
    seen = {}
    for e in entries:
        name = e.get('ingredient_name', '')
        if name and name not in seen:
            seen[name] = {
                'name': name,
                'unit': e.get('unit', ''),
                'latest_cost': e.get('cost_per_unit', 0),
                'latest_supplier': e.get('supplier', ''),
                'last_purchased': e.get('date', '')
            }
    return list(seen.values())


@api_router.put("/inventory/{entry_id}")
async def update_inventory_entry(entry_id: str, input: InventoryEntryCreate):
    total_qty = input.packs_in * input.units_per_pack
    cost_per_unit = (input.pack_cost / input.units_per_pack) if input.units_per_pack > 0 else 0
    result = await db.inventory.update_one({"id": entry_id}, {"$set": {
        "ingredient_name": input.ingredient_name, "date": input.date,
        "packs_in": input.packs_in, "units_per_pack": input.units_per_pack,
        "unit": input.unit, "total_qty_added": round(total_qty, 4),
        "pack_cost": input.pack_cost, "cost_per_unit": round(cost_per_unit, 4),
        "supplier": input.supplier, "notes": input.notes
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Updated"}

@api_router.delete("/inventory/{entry_id}")
async def delete_inventory_entry(entry_id: str):
    result = await db.inventory.delete_one({"id": entry_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"message": "Deleted"}


# -------- CASHFLOW TRACKER --------

@api_router.get("/cashflow")
async def get_cashflow_targets():
    targets = await db.cashflow.find({}, {"_id": 0}).sort("month", 1).to_list(100)
    # Auto-fill actual_sales from sessions
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    month_sales = {}
    for s in sessions:
        m = s.get('date', '')[:7]
        month_sales[m] = month_sales.get(m, 0) + s.get('calculated_sales', 0)
    for t in targets:
        t['actual_sales'] = round(month_sales.get(t['month'], 0), 2)
    return targets

@api_router.post("/cashflow")
async def create_cashflow_target(input: CashflowTargetCreate):
    # Check if month already exists
    existing = await db.cashflow.find_one({"month": input.month}, {"_id": 0})
    if existing:
        await db.cashflow.update_one(
            {"month": input.month},
            {"$set": {
                "sales_target": input.sales_target,
                "growth_saved": input.growth_saved,
                "emergency_saved": input.emergency_saved,
                "notes": input.notes
            }}
        )
        updated = await db.cashflow.find_one({"month": input.month}, {"_id": 0})
        return updated
    target = CashflowTarget(
        month=input.month,
        sales_target=input.sales_target,
        growth_saved=input.growth_saved,
        emergency_saved=input.emergency_saved,
        notes=input.notes
    )
    doc = target.model_dump()
    await db.cashflow.insert_one(doc)
    return serialize_doc(doc)

@api_router.delete("/cashflow/{month}")
async def delete_cashflow_target(month: str):
    result = await db.cashflow.delete_one({"month": month})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Cashflow target not found")
    return {"message": "Deleted"}

# -------- PRODUCT INGREDIENTS (Calculator) --------

@api_router.get("/products/{product_id}/ingredients")
async def get_product_ingredients(product_id: str):
    doc = await db.product_ingredients.find_one({"product_id": product_id}, {"_id": 0})
    if not doc:
        return {"product_id": product_id, "ingredients": []}
    return doc

@api_router.put("/products/{product_id}/ingredients")
async def update_product_ingredients(product_id: str, input: ProductIngredientsUpdate, request: Request):
    await require_owner(request)
    # Calculate total food cost from ingredients
    total_food_cost = 0
    processed = []
    for ing in input.ingredients:
        qty = float(ing.get('qty_per_order', 0))
        cost = float(ing.get('cost_per_unit', 0))
        total = round(qty * cost, 4)
        total_food_cost += total
        processed.append({
            "name": ing.get('name', ''),
            "qty_per_order": qty,
            "unit": ing.get('unit', 'g'),
            "cost_per_unit": cost,
            "total_cost": total
        })

    await db.product_ingredients.update_one(
        {"product_id": product_id},
        {"$set": {"product_id": product_id, "ingredients": processed}},
        upsert=True
    )

    # Auto-update the product's food_cost and derived fields
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if product:
        total_food_cost = round(total_food_cost, 2)
        packaging_cost = product.get('packaging_cost', 0)
        total_cost = total_food_cost + packaging_cost
        price = product.get('price', 0)
        cogs_pct = (total_cost / price * 100) if price > 0 else 0
        profit = price - total_cost
        await db.products.update_one(
            {"id": product_id},
            {"$set": {
                "food_cost": total_food_cost,
                "total_cost": round(total_cost, 2),
                "cogs_percent": round(cogs_pct, 2),
                "profit_per_order": round(profit, 2)
            }}
        )

    return {"product_id": product_id, "ingredients": processed, "calculated_food_cost": round(total_food_cost, 2)}

# -------- EXPORT --------

@api_router.get("/export/sessions")
async def export_sessions_csv():
    from fastapi.responses import StreamingResponse
    import io
    import csv
    sessions = await db.sessions.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Session ID", "Date", "Market", "Total Units", "Calculated Sales", "Cash", "EFTPOS", "Total Collected", "Variance", "Status", "Food COGS", "Packaging", "Total COGS", "Gross Profit", "COGS %"])
    for s in sessions:
        writer.writerow([
            s.get('session_id'), s.get('date'), s.get('market_name'),
            s.get('total_units'), s.get('calculated_sales'), s.get('cash'), s.get('eftpos'),
            s.get('total_collected'), s.get('variance'), s.get('status'),
            s.get('food_cogs'), s.get('packaging'), s.get('total_cogs'),
            s.get('gross_profit'), s.get('cogs_percent')
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=sessions_export.csv"}
    )

@api_router.get("/export/products")
async def export_products_csv():
    from fastapi.responses import StreamingResponse
    import io
    import csv
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name", "Code", "Unit", "Price (NZD)", "Food Cost", "Packaging Cost", "Total Cost", "COGS %", "Profit/Order", "Current Stock", "Reorder Point"])
    for p in products:
        writer.writerow([
            p.get('name'), p.get('code'), p.get('unit'), p.get('price'),
            p.get('food_cost'), p.get('packaging_cost'), p.get('total_cost'),
            round(p.get('cogs_percent', 0), 2), round(p.get('profit_per_order', 0), 2),
            p.get('current_stock'), p.get('reorder_point')
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=products_export.csv"}
    )

# -------- MARKET COMPARISON --------

@api_router.get("/dashboard/market-comparison")
async def get_market_comparison():
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)

    market_data = {}
    for s in sessions:
        mname = s.get('market_name', 'Unknown')
        if mname not in market_data:
            market_data[mname] = {
                'market': mname,
                'sessions': 0, 'total_sales': 0, 'total_cogs': 0,
                'total_profit': 0, 'total_units': 0,
                'product_units': {}
            }
        md = market_data[mname]
        md['sessions'] += 1
        md['total_sales'] += s.get('calculated_sales', 0)
        md['total_cogs'] += s.get('total_cogs', 0)
        md['total_profit'] += s.get('gross_profit', 0)
        md['total_units'] += s.get('total_units', 0)
        for sale in s.get('sales', []):
            pname = sale.get('product_name', '')
            md['product_units'][pname] = md['product_units'].get(pname, 0) + sale.get('units_sold', 0)

    result = []
    for mname, md in market_data.items():
        avg_session_rev = md['total_sales'] / md['sessions'] if md['sessions'] > 0 else 0
        cogs_pct = (md['total_cogs'] / md['total_sales'] * 100) if md['total_sales'] > 0 else 0
        profit_margin = (md['total_profit'] / md['total_sales'] * 100) if md['total_sales'] > 0 else 0
        top_products = sorted(md['product_units'].items(), key=lambda x: x[1], reverse=True)[:3]
        result.append({
            'market': mname,
            'sessions': md['sessions'],
            'total_sales': round(md['total_sales'], 2),
            'total_cogs': round(md['total_cogs'], 2),
            'total_profit': round(md['total_profit'], 2),
            'total_units': md['total_units'],
            'avg_session_revenue': round(avg_session_rev, 2),
            'cogs_percent': round(cogs_pct, 2),
            'profit_margin': round(profit_margin, 2),
            'top_products': [{'name': n, 'units': u} for n, u in top_products]
        })
    result.sort(key=lambda x: x['total_profit'], reverse=True)
    return result

# -------- WEEKLY CONTROL --------

@api_router.get("/dashboard/weekly-control")
async def get_weekly_control():
    sessions = await db.sessions.find({}, {"_id": 0}).sort("date", 1).to_list(1000)
    from datetime import date as dt_date
    week_data = {}
    for s in sessions:
        date_str = s.get('date', '')
        if not date_str:
            continue
        try:
            parts = date_str.split('-')
            d = dt_date(int(parts[0]), int(parts[1]), int(parts[2]))
            iso = d.isocalendar()
            week_key = f"{iso[0]}-W{iso[1]:02d}"
        except (ValueError, TypeError, IndexError):
            continue

        if week_key not in week_data:
            week_data[week_key] = {
                'week': week_key, 'start_date': date_str,
                'sessions': 0, 'sales': 0, 'cash': 0, 'eftpos': 0,
                'total_collected': 0, 'total_cogs': 0, 'gross_profit': 0,
                'expenses': 0, 'total_units': 0, 'markets': set()
            }
        wd = week_data[week_key]
        wd['sessions'] += 1
        wd['sales'] += s.get('calculated_sales', 0)
        wd['cash'] += s.get('cash', 0)
        wd['eftpos'] += s.get('eftpos', 0)
        wd['total_collected'] += s.get('total_collected', 0)
        wd['total_cogs'] += s.get('total_cogs', 0)
        wd['gross_profit'] += s.get('gross_profit', 0)
        wd['expenses'] += s.get('cash_expenses', 0)
        wd['total_units'] += s.get('total_units', 0)
        wd['markets'].add(s.get('market_name', ''))
        wd['end_date'] = date_str

    result = []
    for wk, wd in sorted(week_data.items()):
        net_profit = wd['gross_profit'] - wd['expenses']
        result.append({
            'week': wd['week'],
            'start_date': wd['start_date'],
            'end_date': wd.get('end_date', wd['start_date']),
            'sessions': wd['sessions'],
            'markets': list(wd['markets']),
            'sales': round(wd['sales'], 2),
            'cash': round(wd['cash'], 2),
            'eftpos': round(wd['eftpos'], 2),
            'total_collected': round(wd['total_collected'], 2),
            'total_cogs': round(wd['total_cogs'], 2),
            'gross_profit': round(wd['gross_profit'], 2),
            'expenses': round(wd['expenses'], 2),
            'net_profit': round(net_profit, 2),
            'total_units': wd['total_units'],
            'cogs_percent': round((wd['total_cogs'] / wd['sales'] * 100) if wd['sales'] > 0 else 0, 2)
        })
    return result

# -------- REFILL COST TRENDS --------

@api_router.get("/dashboard/refill-trends")
async def get_refill_trends():
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    inventory = await db.inventory.find({}, {"_id": 0}).sort("date", 1).to_list(1000)
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)

    # Aggregate inventory costs over time per product
    product_trends = {}
    for p in products:
        product_trends[p['id']] = {
            'product_id': p['id'], 'name': p['name'], 'code': p['code'],
            'current_cost': p.get('food_cost', 0),
            'cost_history': [], 'total_spent': 0, 'total_units_bought': 0,
            'lifetime_units_sold': 0, 'lifetime_revenue': 0
        }

    for entry in inventory:
        pid = entry.get('product_id')
        if pid in product_trends:
            cost = entry.get('cost_per_unit', 0)
            qty = entry.get('total_added', 0)
            product_trends[pid]['cost_history'].append({
                'date': entry.get('date', ''),
                'cost_per_unit': cost,
                'units_added': qty,
                'supplier': entry.get('supplier', '')
            })
            product_trends[pid]['total_spent'] += cost * qty
            product_trends[pid]['total_units_bought'] += qty

    for s in sessions:
        for sale in s.get('sales', []):
            pid = sale.get('product_id')
            if pid in product_trends:
                product_trends[pid]['lifetime_units_sold'] += sale.get('units_sold', 0)
                product_trends[pid]['lifetime_revenue'] += sale.get('sales_value', 0)

    result = []
    for pid, pt in product_trends.items():
        avg_cost = (pt['total_spent'] / pt['total_units_bought']) if pt['total_units_bought'] > 0 else pt['current_cost']
        cost_trend = 0
        if len(pt['cost_history']) >= 2:
            first = pt['cost_history'][0]['cost_per_unit']
            last = pt['cost_history'][-1]['cost_per_unit']
            cost_trend = ((last - first) / first * 100) if first > 0 else 0
        result.append({
            **pt,
            'avg_cost': round(avg_cost, 4),
            'cost_trend_pct': round(cost_trend, 1)
        })
    result.sort(key=lambda x: x['lifetime_revenue'], reverse=True)
    return result

# -------- SCALE PLANNER --------

@api_router.get("/dashboard/scale-planner")
async def get_scale_planner(
    target_weekly_revenue: float = 3000,
    weeks_horizon: int = 12
):
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    settings_doc = await db.allocation_settings.find_one({"id": "default"}, {"_id": 0})
    gst_rate = (settings_doc.get('gst_rate', 15) if settings_doc else 15) / 100

    total_sessions = len(sessions)
    weekly = compute_weekly_metrics(sessions)
    weekly["gst_rate"] = gst_rate
    avg_session_rev = weekly["total_sales"] / total_sessions if total_sessions > 0 else 0

    projections = compute_scale_projections(
        target_weekly_revenue, weeks_horizon, avg_session_rev,
        total_sessions, weekly, products
    )

    return {
        'target_weekly_revenue': target_weekly_revenue,
        'weeks_horizon': weeks_horizon,
        'current_weekly_avg': round(weekly["current_weekly"], 2),
        'avg_session_revenue': round(avg_session_rev, 2),
        'avg_cogs_percent': round(weekly["avg_cogs_pct"] * 100, 1),
        'weeks_active': weekly["weeks_active"],
        'total_sessions': total_sessions,
        **projections,
    }


def calc_sales_mix(sessions, products, market_id=None):
    """Calculate historical sales mix percentages per product."""
    product_totals = {}
    total_units = 0
    filtered = [s for s in sessions if not market_id or s.get('market_id') == market_id]
    for s in filtered:
        for sale in s.get('sales', []):
            pid = sale.get('product_id')
            units = sale.get('units_sold', 0)
            product_totals[pid] = product_totals.get(pid, 0) + units
            total_units += units
    return product_totals, total_units

def build_checklist_items(products, product_totals, total_units, target_revenue):
    """Build prep checklist items from products and sales mix."""
    checklist = []
    total_est_cost = 0
    for p in products:
        mix = (product_totals.get(p['id'], 0) / total_units * 100) if total_units > 0 else 10
        est_orders = int((target_revenue * (mix / 100)) / p['price']) if p['price'] > 0 else 0
        current = p.get('current_stock', 0)
        gap = max(0, est_orders - current)
        cost = gap * p.get('total_cost', 0)
        total_est_cost += cost
        checklist.append({
            'product_id': p['id'], 'product_name': p['name'], 'code': p['code'],
            'estimated_orders': est_orders, 'current_stock': current, 'to_prep': gap,
            'unit_cost': p.get('total_cost', 0), 'estimated_cost': round(cost, 2),
            'status': 'Ready' if gap == 0 else 'Prep Needed', 'checked': False
        })
    return checklist, total_est_cost

def accumulate_staff_session(staff_stats, session):
    """Accumulate a single session into staff stats."""
    uid = session.get('created_by_id', '')
    uname = session.get('created_by_name', '') or 'System (Seeded)'
    urole = session.get('created_by_role', '') or 'system'
    key = uid or 'system'

    if key not in staff_stats:
        staff_stats[key] = {
            'user_id': uid, 'name': uname, 'role': urole,
            'sessions': 0, 'total_sales': 0, 'total_profit': 0,
            'total_units': 0, 'total_cogs': 0, 'markets': set(),
            'best_session_sales': 0, 'best_session_date': '', 'dates': []
        }

    st = staff_stats[key]
    st['sessions'] += 1
    st['total_sales'] += session.get('calculated_sales', 0)
    st['total_profit'] += session.get('gross_profit', 0)
    st['total_units'] += session.get('total_units', 0)
    st['total_cogs'] += session.get('total_cogs', 0)
    st['markets'].add(session.get('market_name', ''))
    st['dates'].append(session.get('date', ''))
    if session.get('calculated_sales', 0) > st['best_session_sales']:
        st['best_session_sales'] = session.get('calculated_sales', 0)
        st['best_session_date'] = session.get('date', '')

# -------- DAILY PREP CHECKLIST --------

@api_router.get("/prep-checklist")
async def generate_prep_checklist(market_id: Optional[str] = None, target_revenue: float = 1000):
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    markets = await db.markets.find({}, {"_id": 0}).to_list(100)

    market_name = "All Markets"
    if market_id:
        m = next((mk for mk in markets if mk['id'] == market_id), None)
        if m:
            market_name = m['name']

    product_totals, total_units = calc_sales_mix(sessions, products, market_id)
    checklist, total_est_cost = build_checklist_items(products, product_totals, total_units, target_revenue)

    return {
        'date': datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        'market_name': market_name, 'market_id': market_id,
        'target_revenue': target_revenue, 'checklist': checklist,
        'total_items_to_prep': sum(1 for c in checklist if c['to_prep'] > 0),
        'total_estimated_cost': round(total_est_cost, 2),
        'markets': [{'id': mk['id'], 'name': mk['name']} for mk in markets]
    }

@api_router.get("/export/prep-checklist")
async def export_prep_checklist_csv(market_id: Optional[str] = None, target_revenue: float = 1000):
    from fastapi.responses import StreamingResponse
    import io
    import csv
    data = await generate_prep_checklist(market_id, target_revenue)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["GRILL SHACK - Daily Prep Checklist"])
    writer.writerow([f"Market: {data['market_name']}", f"Date: {data['date']}", f"Target: ${data['target_revenue']}"])
    writer.writerow([])
    writer.writerow(["Product", "Code", "Est. Orders", "In Stock", "To Prep", "Unit Cost", "Est. Cost", "Status", "Done?"])
    for item in data['checklist']:
        writer.writerow([item['product_name'], item['code'], item['estimated_orders'], item['current_stock'], item['to_prep'], f"${item['unit_cost']:.2f}", f"${item['estimated_cost']:.2f}", item['status'], "[ ]"])
    writer.writerow([])
    writer.writerow([f"Total items to prep: {data['total_items_to_prep']}", f"Estimated cost: ${data['total_estimated_cost']:.2f}"])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=prep_checklist_{data['date']}.csv"})

# -------- ALERTS --------

@api_router.get("/alerts")
async def get_alerts():
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    sessions = await db.sessions.find({}, {"_id": 0}).sort("date", -1).to_list(20)
    alerts = []
    for p in products:
        if p.get('current_stock', 0) <= p.get('reorder_point', 10):
            severity = 'critical' if p.get('current_stock', 0) == 0 else 'warning'
            alerts.append({'id': f"stock-{p['id']}", 'type': 'stock', 'severity': severity, 'title': f"{p['name']} - {'Out of Stock!' if severity == 'critical' else 'Low Stock'}", 'message': f"Current: {p.get('current_stock', 0)} units, Reorder at: {p.get('reorder_point', 10)}", 'product_id': p['id'], 'product_code': p['code']})
    for s in sessions[:5]:
        if abs(s.get('variance', 0)) > 50:
            alerts.append({'id': f"cash-{s['id']}", 'type': 'cash', 'severity': 'warning', 'title': f"Large variance: Session #{s.get('session_id')}", 'message': f"Variance of ${abs(s.get('variance', 0)):.2f} at {s.get('market_name')} on {s.get('date')}", 'session_id': s['id']})
    for p in products:
        if p.get('cogs_percent', 0) > 40:
            alerts.append({'id': f"cogs-{p['id']}", 'type': 'cogs', 'severity': 'info', 'title': f"High COGS: {p['name']}", 'message': f"COGS at {p.get('cogs_percent', 0):.1f}% - review pricing", 'product_id': p['id']})
    alerts.sort(key=lambda x: {'critical': 0, 'warning': 1, 'info': 2}.get(x['severity'], 3))
    return alerts



# -------- STAFF PERFORMANCE --------

@api_router.get("/dashboard/staff-performance")
async def get_staff_performance():
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)

    staff_stats = {}
    for s in sessions:
        accumulate_staff_session(staff_stats, s)

    result = []
    for key, st in staff_stats.items():
        avg_session = st['total_sales'] / st['sessions'] if st['sessions'] > 0 else 0
        cogs_pct = (st['total_cogs'] / st['total_sales'] * 100) if st['total_sales'] > 0 else 0
        result.append({
            'user_id': st['user_id'],
            'name': st['name'],
            'role': st['role'],
            'sessions': st['sessions'],
            'total_sales': round(st['total_sales'], 2),
            'total_profit': round(st['total_profit'], 2),
            'total_units': st['total_units'],
            'total_cogs': round(st['total_cogs'], 2),
            'avg_session_revenue': round(avg_session, 2),
            'cogs_percent': round(cogs_pct, 2),
            'markets_worked': list(st['markets']),
            'best_session_sales': round(st['best_session_sales'], 2),
            'best_session_date': st['best_session_date'],
            'first_session': min(st['dates']) if st['dates'] else '',
            'last_session': max(st['dates']) if st['dates'] else ''
        })
    result.sort(key=lambda x: x['total_sales'], reverse=True)
    return result


# -------- AUTH ENDPOINTS --------

class LoginInput(BaseModel):
    email: str
    password: str

class RegisterInput(BaseModel):
    email: str
    password: str
    name: str
    role: str = "staff"

@api_router.post("/auth/login")
async def login(input: LoginInput):
    email = input.email.strip().lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not verify_password(input.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    token = create_access_token(user["id"], email, user.get("role", "staff"))
    refresh = create_refresh_token(user["id"])
    response = JSONResponse(content={
        "id": user["id"], "email": user["email"], "name": user.get("name", ""),
        "role": user.get("role", "staff"), "token": token
    })
    response.set_cookie(key="access_token", value=token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    return response

@api_router.post("/auth/register")
async def register(input: RegisterInput):
    email = input.email.strip().lower()
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    user_id = str(uuid.uuid4())
    await db.users.insert_one({
        "id": user_id, "email": email, "name": input.name,
        "password_hash": hash_password(input.password),
        "role": input.role if input.role in ["owner", "staff"] else "staff",
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    token = create_access_token(user_id, email, input.role)
    refresh = create_refresh_token(user_id)
    response = JSONResponse(content={
        "id": user_id, "email": email, "name": input.name, "role": input.role, "token": token
    })
    response.set_cookie(key="access_token", value=token, httponly=True, secure=False, samesite="lax", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    return response

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    return user

@api_router.post("/auth/logout")
async def logout():
    response = JSONResponse(content={"message": "Logged out"})
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return response

@api_router.get("/auth/users")
async def get_users(request: Request):
    user = await get_current_user(request)
    if user.get("role") != "owner":
        raise HTTPException(status_code=403, detail="Owner only")
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(100)
    return users

# -------- SUPPLIER DIRECTORY --------

class SupplierCreate(BaseModel):
    name: str
    contact_person: str = ""
    phone: str = ""
    email: str = ""
    address: str = ""
    products: List[str] = []
    notes: str = ""

@api_router.get("/suppliers")
async def get_suppliers():
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(100)
    return suppliers

@api_router.post("/suppliers")
async def create_supplier(input: SupplierCreate, request: Request):
    await require_owner(request)
    doc = {
        "id": str(uuid.uuid4()), "name": input.name,
        "contact_person": input.contact_person, "phone": input.phone,
        "email": input.email, "address": input.address,
        "products": input.products, "notes": input.notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.suppliers.insert_one(doc)
    doc.pop('_id', None)
    return doc

@api_router.put("/suppliers/{supplier_id}")
async def update_supplier(supplier_id: str, input: SupplierCreate, request: Request):
    await require_owner(request)
    result = await db.suppliers.update_one({"id": supplier_id}, {"$set": {
        "name": input.name, "contact_person": input.contact_person,
        "phone": input.phone, "email": input.email,
        "address": input.address, "products": input.products, "notes": input.notes
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Updated"}

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, request: Request):
    await require_owner(request)
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Deleted"}

# -------- AUTO-REORDER / PURCHASE ORDERS --------

class PurchaseOrderItem(BaseModel):
    product_id: str
    product_name: str
    qty_needed: int
    unit_cost: float = 0
    estimated_cost: float = 0

class PurchaseOrderCreate(BaseModel):
    supplier_id: str
    supplier_name: str
    items: List[dict]
    notes: str = ""

@api_router.get("/reorder/suggestions")
async def get_reorder_suggestions(request: Request):
    await require_owner(request)
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(100)

    suggestions = []
    for p in products:
        current = p.get('current_stock', 0)
        reorder_pt = p.get('reorder_point', 0)
        if current < reorder_pt:
            qty_needed = max(reorder_pt * 2 - current, reorder_pt)
            unit_cost = p.get('food_cost', 0) + p.get('packaging_cost', 0)
            # Find linked supplier
            linked_supplier = None
            for s in suppliers:
                if p.get('code') in (s.get('products') or []):
                    linked_supplier = {"id": s['id'], "name": s['name']}
                    break
            suggestions.append({
                "product_id": p['id'],
                "product_name": p['name'],
                "code": p['code'],
                "current_stock": current,
                "reorder_point": reorder_pt,
                "qty_needed": qty_needed,
                "unit_cost": round(unit_cost, 2),
                "estimated_cost": round(qty_needed * unit_cost, 2),
                "linked_supplier": linked_supplier,
            })
    return suggestions

@api_router.post("/reorder/purchase-orders")
async def create_purchase_order(input: PurchaseOrderCreate, request: Request):
    user = await require_owner(request)
    total = sum(item.get('estimated_cost', 0) for item in input.items)
    po = {
        "id": str(uuid.uuid4()),
        "po_number": f"PO-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M')}",
        "supplier_id": input.supplier_id,
        "supplier_name": input.supplier_name,
        "items": input.items,
        "total_estimated": round(total, 2),
        "status": "pending",
        "notes": input.notes,
        "created_by": user.get("name", user.get("email", "")),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.purchase_orders.insert_one(po)
    po.pop('_id', None)
    return po

@api_router.get("/reorder/purchase-orders")
async def get_purchase_orders(request: Request, status: Optional[str] = None):
    await require_owner(request)
    query = {}
    if status:
        query["status"] = status
    orders = await db.purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return orders

@api_router.put("/reorder/purchase-orders/{po_id}/status")
async def update_po_status(po_id: str, request: Request, new_status: str = Query(...)):
    await require_owner(request)
    if new_status not in ("pending", "ordered", "received", "cancelled"):
        raise HTTPException(status_code=400, detail="Invalid status")
    result = await db.purchase_orders.update_one(
        {"id": po_id},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    return {"message": f"Status updated to {new_status}"}

@api_router.delete("/reorder/purchase-orders/{po_id}")
async def delete_purchase_order(po_id: str, request: Request):
    await require_owner(request)
    result = await db.purchase_orders.delete_one({"id": po_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    return {"message": "Deleted"}

# -------- HISTORICAL COMPARISON --------

@api_router.get("/dashboard/historical")
async def get_historical_comparison():
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(1000)
    from datetime import date as dt_date

    def week_key(s):
        try:
            parts = s.get('date', '').split('-')
            d = dt_date(int(parts[0]), int(parts[1]), int(parts[2]))
            iso = d.isocalendar()
            return f"{iso[0]}-W{iso[1]:02d}"
        except (ValueError, IndexError):
            return None

    def month_key(s):
        return s.get('date', '')[:7] or None

    def calc_growth_series(period_data):
        sorted_keys = sorted(period_data.keys())
        result = []
        for i, key in enumerate(sorted_keys):
            d = period_data[key]
            prev = period_data[sorted_keys[i - 1]] if i > 0 else None
            growth = ((d['sales'] - prev['sales']) / prev['sales'] * 100) if prev and prev['sales'] > 0 else 0
            # Add human-readable label
            label = key
            if key.startswith("20") and "-W" in key:
                # Week format: 2026-W08 → "2026 Week 8 (Feb)"
                parts = key.split("-W")
                yr, wk = parts[0], int(parts[1])
                try:
                    from datetime import date as dt_date2
                    d_ref = dt_date2.fromisocalendar(int(yr), wk, 1)
                    month_name = d_ref.strftime('%b')
                    label = f"{yr} Week {wk} ({month_name})"
                except (ValueError, TypeError):
                    label = f"{yr} Week {wk}"
            elif key.startswith("20") and len(key) == 7:
                # Month format: 2026-03 → "March 2026"
                try:
                    from datetime import date as dt_date2
                    d_ref = dt_date2(int(key[:4]), int(key[5:7]), 1)
                    label = d_ref.strftime('%B %Y')
                except (ValueError, TypeError):
                    pass
            result.append({
                'period': key, 'label': label,
                'sales': round(d['sales'], 2), 'profit': round(d['profit'], 2),
                'units': d['units'], 'sessions': d['sessions'],
                'cogs': round(d['cogs'], 2), 'growth_pct': round(growth, 1)
            })
        return result

    week_data = aggregate_by_period(sessions, week_key)
    month_data = aggregate_by_period(sessions, month_key)

    return {
        "week_over_week": calc_growth_series(week_data),
        "month_over_month": calc_growth_series(month_data)
    }



# -------- MARKET PRESET COPY --------

@api_router.post("/markets/{market_id}/copy-preset")
async def copy_market_preset(market_id: str, request: Request, source_market_id: str = Query(...)):
    await require_owner(request)
    source = await db.markets.find_one({"id": source_market_id}, {"_id": 0})
    if not source:
        raise HTTPException(status_code=404, detail="Source market not found")
    target = await db.markets.find_one({"id": market_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="Target market not found")
    await db.markets.update_one({"id": market_id}, {"$set": {"preset_mix": source.get("preset_mix", {})}})
    return {"message": f"Preset copied from {source['name']} to {target['name']}"}

# -------- EXCEL EXPORT --------

from fastapi.responses import StreamingResponse
import io

def _make_excel_workbook(title, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = title
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    for col in ws.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    return wb

@api_router.get("/export/sessions-excel")
async def export_sessions_excel():
    sessions = await db.sessions.find({}, {"_id": 0}).sort("date", -1).to_list(5000)
    headers = ["Date", "Market", "Units", "Calculated Sales", "COGS", "Gross Profit", "Cash", "EFTPOS", "Total Collected", "Variance", "Status", "Cash Expenses", "Notes"]
    rows = []
    for s in sessions:
        rows.append([s.get('date'), s.get('market_name'), s.get('total_units', 0), s.get('calculated_sales', 0), s.get('total_cogs', 0), s.get('gross_profit', 0), s.get('cash', 0), s.get('eftpos', 0), s.get('total_collected', 0), s.get('variance', 0), s.get('status', ''), s.get('cash_expenses', 0), s.get('notes', '')])
    wb = _make_excel_workbook("Sessions", headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=GrillShack_Sessions.xlsx"})

@api_router.get("/export/products-excel")
async def export_products_excel():
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    headers = ["Name", "Code", "Price", "Food Cost", "Packaging Cost", "Total Cost", "COGS %", "Profit/Unit", "Current Stock", "Reorder Point"]
    rows = [[p.get('name'), p.get('code'), p.get('price', 0), p.get('food_cost', 0), p.get('packaging_cost', 0), p.get('total_cost', 0), p.get('cogs_percent', 0), p.get('profit_per_order', 0), p.get('current_stock', 0), p.get('reorder_point', 0)] for p in products]
    wb = _make_excel_workbook("Products", headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=GrillShack_Products.xlsx"})

@api_router.get("/export/inventory-excel")
async def export_inventory_excel():
    entries = await db.inventory.find({}, {"_id": 0}).sort("date", -1).to_list(5000)
    headers = ["Date", "Ingredient", "Packs", "Units/Pack", "Unit", "Pack Cost", "Cost/Unit", "Total Qty", "Total Cost", "Supplier", "Notes"]
    rows = [[e.get('date'), e.get('ingredient_name'), e.get('packs_in', 0), e.get('units_per_pack', 0), e.get('unit', ''), e.get('pack_cost', 0), e.get('cost_per_unit', 0), e.get('total_qty_added', 0), e.get('total_cost', 0), e.get('supplier', ''), e.get('notes', '')] for e in entries]
    wb = _make_excel_workbook("Inventory", headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=GrillShack_Inventory.xlsx"})

@api_router.get("/export/cashflow-excel")
async def export_cashflow_excel():
    targets = await db.cashflow_targets.find({}, {"_id": 0}).to_list(500)
    headers = ["Month", "Sales Target", "Actual Sales", "Growth Saved", "Emergency Saved", "Notes"]
    rows = [[t.get('month'), t.get('sales_target', 0), t.get('actual_sales', 0), t.get('growth_saved', 0), t.get('emergency_saved', 0), t.get('notes', '')] for t in targets]
    wb = _make_excel_workbook("Cashflow", headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=GrillShack_Cashflow.xlsx"})

@api_router.get("/export/sales-by-month-market")
async def export_sales_by_month_market():
    """Sales by month with all markets breakdown — Excel download."""
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(5000)
    markets_set = sorted(set(s.get('market_name', 'Unknown') for s in sessions))
    data = {}
    for s in sessions:
        month = s.get('date', '')[:7]
        market = s.get('market_name', 'Unknown')
        if month not in data:
            data[month] = {m: {'sales': 0, 'units': 0, 'profit': 0} for m in markets_set}
        data[month][market]['sales'] += s.get('calculated_sales', 0)
        data[month][market]['units'] += s.get('total_units', 0)
        data[month][market]['profit'] += s.get('gross_profit', 0)
    headers = ["Month"] + [f"{m} Sales" for m in markets_set] + [f"{m} Units" for m in markets_set] + ["Total Sales", "Total Units", "Total Profit"]
    rows = []
    for month in sorted(data.keys()):
        row = [month]
        total_s, total_u, total_p = 0, 0, 0
        for m in markets_set:
            row.append(round(data[month][m]['sales'], 2))
            total_s += data[month][m]['sales']
        for m in markets_set:
            row.append(data[month][m]['units'])
            total_u += data[month][m]['units']
        total_p = sum(data[month][m]['profit'] for m in markets_set)
        row.extend([round(total_s, 2), total_u, round(total_p, 2)])
        rows.append(row)
    wb = _make_excel_workbook("Sales by Month & Market", headers, rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=GrillShack_Sales_By_Month_Market.xlsx"})

# -------- DATA BACKUP / RESTORE --------

import json as json_module

@api_router.get("/backup/export")
async def export_backup(request: Request):
    """Export all data as a JSON backup (safe mode)."""
    await require_owner(request)
    collections = ['products', 'markets', 'sessions', 'inventory', 'cashflow_targets',
                    'allocation_settings', 'suppliers', 'purchase_orders', 'users',
                    'product_ingredients', 'data_snapshots']
    backup = {"backup_date": datetime.now(timezone.utc).isoformat(), "version": "2.0"}
    for coll_name in collections:
        coll = db[coll_name]
        docs = await coll.find({}, {"_id": 0}).to_list(10000)
        backup[coll_name] = docs
    content = json_module.dumps(backup, indent=2, default=str)
    return StreamingResponse(
        io.BytesIO(content.encode()),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=GrillShack_Backup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.json"}
    )

@api_router.post("/backup/restore")
async def restore_backup(request: Request):
    """Restore data from a JSON backup. Owner-only."""
    await require_owner(request)
    body = await request.json()
    if "version" not in body or "products" not in body:
        raise HTTPException(status_code=400, detail="Invalid backup file format")
    collections = ['products', 'markets', 'sessions', 'inventory', 'cashflow_targets',
                    'allocation_settings', 'suppliers', 'purchase_orders',
                    'product_ingredients', 'data_snapshots']
    restored = {}
    for coll_name in collections:
        if coll_name in body and body[coll_name]:
            coll = db[coll_name]
            await coll.delete_many({})
            await coll.insert_many(body[coll_name])
            restored[coll_name] = len(body[coll_name])
    return {"message": "Backup restored successfully", "restored": restored}

# -------- DATA SNAPSHOTS (versioning for stock planner lookback) --------

@api_router.post("/snapshots")
async def create_snapshot(request: Request):
    """Save a snapshot of current data for Stock Planner lookback. Owner-only."""
    await require_owner(request)
    body = await request.json()
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    sessions = await db.sessions.find({}, {"_id": 0}).to_list(5000)
    snapshot = {
        "id": str(uuid.uuid4()),
        "label": body.get("label", f"Snapshot {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}"),
        "notes": body.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "products": products,
        "session_summary": {
            "count": len(sessions),
            "total_sales": round(sum(s.get('calculated_sales', 0) for s in sessions), 2),
            "total_profit": round(sum(s.get('gross_profit', 0) for s in sessions), 2),
            "total_cogs": round(sum(s.get('total_cogs', 0) for s in sessions), 2),
        }
    }
    await db.data_snapshots.insert_one(snapshot)
    snapshot.pop('_id', None)
    return snapshot

@api_router.get("/snapshots")
async def list_snapshots(request: Request):
    await require_owner(request)
    snapshots = await db.data_snapshots.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return snapshots

@api_router.get("/snapshots/{snapshot_id}")
async def get_snapshot(snapshot_id: str, request: Request):
    await require_owner(request)
    doc = await db.data_snapshots.find_one({"id": snapshot_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Snapshot not found")
    return doc

@api_router.put("/snapshots/{snapshot_id}")
async def update_snapshot(snapshot_id: str, request: Request):
    """Edit snapshot label/notes (for rectification). Owner-only."""
    await require_owner(request)
    body = await request.json()
    update = {}
    if "label" in body:
        update["label"] = body["label"]
    if "notes" in body:
        update["notes"] = body["notes"]
    if "products" in body:
        update["products"] = body["products"]
    if not update:
        raise HTTPException(status_code=400, detail="Nothing to update")
    result = await db.data_snapshots.update_one({"id": snapshot_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Snapshot not found")
    return {"message": "Snapshot updated"}

@api_router.delete("/snapshots/{snapshot_id}")
async def delete_snapshot(snapshot_id: str, request: Request):
    await require_owner(request)
    result = await db.data_snapshots.delete_one({"id": snapshot_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Snapshot not found")
    return {"message": "Snapshot deleted"}

# -------- PASSWORD CHANGE --------

class PasswordChangeInput(BaseModel):
    current_password: str
    new_password: str

@api_router.post("/auth/change-password")
async def change_password(input: PasswordChangeInput, request: Request):
    user = await get_current_user(request)
    full_user = await db.users.find_one({"id": user["id"]})
    if not full_user or not verify_password(input.current_password, full_user.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    if len(input.new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    new_hash = hash_password(input.new_password)
    await db.users.update_one({"id": user["id"]}, {"$set": {"password_hash": new_hash}})
    return {"message": "Password changed successfully"}


# -------- SEED DATA --------

async def _seed_products():
    """Seed product data from Excel."""
    products_data = [
        {"name": "Brekkie Buns", "code": "BB", "price": 10.5, "food_cost": 2.49, "packaging_cost": 0.50, "opening_stock": 25},
        {"name": "BBQ Ribs", "code": "BR", "price": 20.5, "food_cost": 5.26, "packaging_cost": 0.50, "opening_stock": 55},
        {"name": "Wagyu Beef Sliders", "code": "WS", "price": 15.5, "food_cost": 2.46, "packaging_cost": 0.50, "opening_stock": 13},
        {"name": "Smokies", "code": "SM", "price": 12.5, "food_cost": 3.88, "packaging_cost": 0.50, "opening_stock": 12},
        {"name": "Pulled Pork", "code": "PP", "price": 15.5, "food_cost": 1.86, "packaging_cost": 0.50, "opening_stock": 3},
        {"name": "Wedges", "code": "WD", "price": 9.5, "food_cost": 1.43, "packaging_cost": 0.50, "opening_stock": 10},
        {"name": "Onion Rings", "code": "OR", "price": 9.5, "food_cost": 2.08, "packaging_cost": 0.50, "opening_stock": 7},
        {"name": "Chicken Lollies", "code": "CP", "price": 8.5, "food_cost": 2.69, "packaging_cost": 0.50, "opening_stock": 20},
        {"name": "Drinks", "code": "DR", "price": 5.0, "food_cost": 0, "packaging_cost": 0, "opening_stock": 47},
        {"name": "Up-sell", "code": "US", "price": 7.0, "food_cost": 1.43, "packaging_cost": 0.50, "opening_stock": 30},
    ]
    for p_data in products_data:
        product = Product(**p_data, current_stock=p_data["opening_stock"])
        doc = product.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.products.insert_one(doc)
    return len(products_data)

async def _seed_markets():
    """Seed market data from Excel."""
    markets_data = [
        {"name": "Grafton Market", "preset_mix": {"BB": 11.4, "BR": 9.0, "WS": 12.1, "SM": 7.2, "PP": 8.5, "WD": 6.6, "OR": 40.0, "CP": 12.1, "DR": 27.9, "US": 31.6}},
        {"name": "Britomart Market", "preset_mix": {"BB": 11.6, "BR": 10.0, "WS": 10.2, "SM": 6.5, "PP": 13.0, "WD": 18.8, "OR": 10.8, "CP": 7.0, "DR": 6.5, "US": 6.8}},
        {"name": "EventFinda Market", "preset_mix": {"BB": 5.0, "BR": 12.0, "WS": 22.0, "SM": 19.0, "PP": 9.0, "WD": 4.0, "OR": 5.0, "CP": 2.0, "DR": 2.0, "US": 2.0}},
        {"name": "Victoria Park Market", "preset_mix": {"BB": 3.0, "BR": 15.0, "WS": 25.0, "SM": 9.0, "PP": 14.0, "WD": 1.0, "OR": 4.0, "CP": 8.0, "DR": 5.0, "US": 6.0}},
        {"name": "Long Bay Market", "preset_mix": {}},
    ]
    markets_map = {}
    for m_data in markets_data:
        market = Market(name=m_data["name"], preset_mix=m_data["preset_mix"])
        await db.markets.insert_one(market.model_dump())
        markets_map[m_data["name"]] = market.id
    return markets_map, len(markets_data)

async def _seed_sessions(markets_map):
    """Seed historical session data from Excel."""
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    products_map = {p['code']: p['id'] for p in products}
    sessions_data = [
        {"date": "2026-02-15", "market": "Grafton Market", "cash": 300, "eftpos": 294.96, "sales": {"BB": 0, "BR": 15, "WS": 10, "SM": 6, "PP": 4, "WD": 4, "OR": 4, "CP": 9, "DR": 0, "US": 0}},
        {"date": "2026-02-20", "market": "Britomart Market", "cash": 0, "eftpos": 0, "sales": {"BB": 10, "BR": 9, "WS": 21, "SM": 13, "PP": 10, "WD": 4, "OR": 4, "CP": 0, "DR": 0, "US": 0}},
        {"date": "2026-03-05", "market": "EventFinda Market", "cash": 0, "eftpos": 0, "sales": {"BB": 0, "BR": 13, "WS": 24, "SM": 21, "PP": 10, "WD": 4, "OR": 5, "CP": 0, "DR": 0, "US": 0}},
        {"date": "2026-03-08", "market": "Britomart Market", "cash": 0, "eftpos": 0, "sales": {"BB": 18, "BR": 11, "WS": 21, "SM": 15, "PP": 9, "WD": 9, "OR": 3, "CP": 0, "DR": 0, "US": 0}},
        {"date": "2026-03-12", "market": "Britomart Market", "cash": 0, "eftpos": 0, "sales": {"BB": 12, "BR": 6, "WS": 17, "SM": 12, "PP": 7, "WD": 1, "OR": 5, "CP": 8, "DR": 0, "US": 0}},
        {"date": "2026-03-15", "market": "Britomart Market", "cash": 500, "eftpos": 660, "sales": {"BB": 16, "BR": 8, "WS": 21, "SM": 15, "PP": 7, "WD": 5, "OR": 1, "CP": 8, "DR": 11, "US": 6}},
        {"date": "2026-03-18", "market": "Victoria Park Market", "cash": 700, "eftpos": 830.4, "sales": {"BB": 0, "BR": 19, "WS": 32, "SM": 11, "PP": 18, "WD": 1, "OR": 5, "CP": 10, "DR": 7, "US": 8}},
        {"date": "2026-03-22", "market": "Britomart Market", "cash": 600, "eftpos": 876.58, "sales": {"BB": 10, "BR": 18, "WS": 23, "SM": 13, "PP": 8, "WD": 8, "OR": 3, "CP": 2, "DR": 4, "US": 6}},
        {"date": "2026-04-02", "market": "Britomart Market", "cash": 550, "eftpos": 819.81, "sales": {"BB": 19, "BR": 7, "WS": 32, "SM": 15, "PP": 6, "WD": 3, "OR": 3, "CP": 9, "DR": 15, "US": 3}},
        {"date": "2026-04-05", "market": "Grafton Market", "cash": 300, "eftpos": 294.96, "sales": {"BB": 0, "BR": 9, "WS": 11, "SM": 4, "PP": 1, "WD": 0, "OR": 1, "CP": 8, "DR": 11, "US": 6}},
        {"date": "2026-04-08", "market": "Britomart Market", "cash": 450, "eftpos": 620.35, "sales": {"BB": 8, "BR": 12, "WS": 19, "SM": 14, "PP": 6, "WD": 3, "OR": 3, "CP": 3, "DR": 10, "US": 7}},
    ]
    for s_data in sessions_data:
        market_id = markets_map.get(s_data["market"], "")
        sales_list = [{"product_id": products_map[code], "units_sold": units} for code, units in s_data["sales"].items() if units > 0 and code in products_map]
        await create_session(SessionCreate(date=s_data["date"], market_id=market_id, market_name=s_data["market"], cash=s_data["cash"], eftpos=s_data["eftpos"], sales=sales_list))
    return len(sessions_data)

@api_router.post("/seed")
async def seed_data():
    """Seed initial data from the Excel file"""
    existing = await db.products.count_documents({})
    if existing > 0:
        return {"message": "Data already seeded", "products": existing}
    
    n_products = await _seed_products()
    markets_map, n_markets = await _seed_markets()
    n_sessions = await _seed_sessions(markets_map)
    await db.allocation_settings.insert_one(AllocationSettings().model_dump())
    
    return {"message": "Data seeded successfully", "products": n_products, "markets": n_markets, "sessions": n_sessions}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    await seed_admin()
    await db.users.create_index("email", unique=True)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
